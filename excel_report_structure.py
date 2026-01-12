"""
Blinkit Sales Performance Analytics - Excel Report Generation
Creates management-ready Excel reports with pivot tables and conditional formatting
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule
import warnings
warnings.filterwarnings('ignore')

print("="*70)
print("GENERATING EXCEL MANAGEMENT REPORT")
print("="*70)

# Load cleaned data
df = pd.read_csv('blinkit_master_data.csv')
df_detailed = pd.read_csv('blinkit_detailed_data.csv')

# Ensure datetime
df['order_date'] = pd.to_datetime(df['order_date'])
df['month'] = df['order_date'].dt.to_period('M').astype(str)

# Create Excel writer
excel_file = 'Blinkit_Management_Report.xlsx'
writer = pd.ExcelWriter(excel_file, engine='openpyxl')

# ============================================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================================
print("\n[1] Creating Executive Summary...")

summary_data = {
    'Metric': [
        'Total Revenue (₹)',
        'Total Profit (₹)',
        'Overall Profit Margin (%)',
        'Total Orders',
        'Delivered Orders',
        'Cancelled Orders (%)',
        'Avg Order Value (₹)',
        'Avg Delivery Time (min)',
        'SLA Breach Rate (%)',
        'Repeat Customer Rate (%)',
        'Avg Discount (%)'
    ],
    'Value': [
        f"{df['revenue'].sum():,.0f}",
        f"{df['profit'].sum():,.0f}",
        f"{(df['profit'].sum() / df['revenue'].sum() * 100):.2f}",
        f"{len(df):,}",
        f"{len(df[df['order_status']=='Delivered']):,}",
        f"{(len(df[df['order_status']=='Cancelled']) / len(df) * 100):.2f}",
        f"{df['revenue'].mean():,.2f}",
        f"{df['delivery_time_minutes'].mean():.1f}",
        f"{(df['delivery_sla_breach'].mean() * 100):.2f}",
        f"{(df['repeat_customer_flag'].mean() * 100):.2f}",
        f"{((df['discount_amount'] / (df['revenue'] + df['discount_amount'])) * 100).mean():.2f}"
    ]
}
summary_df = pd.DataFrame(summary_data)
summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)

# ============================================================================
# SHEET 2: CITY PERFORMANCE
# ============================================================================
print("[2] Creating City Performance Analysis...")

city_perf = df.groupby('city').agg({
    'revenue': 'sum',
    'profit': 'sum',
    'order_id': 'count',
    'delivery_time_minutes': 'mean',
    'delivery_sla_breach': 'mean',
    'repeat_customer_flag': 'mean'
}).round(2)

city_perf.columns = ['Revenue', 'Profit', 'Orders', 'Avg_Delivery_Time', 
                     'SLA_Breach_Rate', 'Repeat_Customer_Rate']
city_perf['Profit_Margin_%'] = (city_perf['Profit'] / city_perf['Revenue'] * 100).round(2)
city_perf['Revenue_Per_Order'] = (city_perf['Revenue'] / city_perf['Orders']).round(2)
city_perf = city_perf.sort_values('Profit', ascending=False).reset_index()

city_perf.to_excel(writer, sheet_name='City Performance', index=False)

# ============================================================================
# SHEET 3: CATEGORY ANALYSIS
# ============================================================================
print("[3] Creating Category Analysis...")

category_perf = df_detailed.groupby('category').agg({
    'selling_price': 'sum',
    'profit': 'sum',
    'order_id': 'nunique'
}).round(2)

category_perf.columns = ['Revenue', 'Profit', 'Orders']
category_perf['Profit_Margin_%'] = (category_perf['Profit'] / category_perf['Revenue'] * 100).round(2)
category_perf['Avg_Order_Value'] = (category_perf['Revenue'] / category_perf['Orders']).round(2)
category_perf = category_perf.sort_values('Revenue', ascending=False).reset_index()

category_perf.to_excel(writer, sheet_name='Category Analysis', index=False)

# ============================================================================
# SHEET 4: MONTHLY TRENDS
# ============================================================================
print("[4] Creating Monthly Trends...")

monthly_trends = df.groupby('month').agg({
    'revenue': 'sum',
    'profit': 'sum',
    'order_id': 'count',
    'delivery_time_minutes': 'mean'
}).round(2)

monthly_trends.columns = ['Revenue', 'Profit', 'Orders', 'Avg_Delivery_Time']
monthly_trends['Profit_Margin_%'] = (monthly_trends['Profit'] / monthly_trends['Revenue'] * 100).round(2)
monthly_trends['Revenue_Growth_%'] = monthly_trends['Revenue'].pct_change() * 100
monthly_trends['Order_Growth_%'] = monthly_trends['Orders'].pct_change() * 100
monthly_trends = monthly_trends.round(2).reset_index()

monthly_trends.to_excel(writer, sheet_name='Monthly Trends', index=False)

# ============================================================================
# SHEET 5: DISCOUNT ANALYSIS
# ============================================================================
print("[5] Creating Discount Analysis...")

df['discount_pct'] = (df['discount_amount'] / (df['revenue'] + df['discount_amount']) * 100).round(2)
df['discount_bucket'] = pd.cut(df['discount_pct'], 
                                bins=[0, 5, 10, 15, 100], 
                                labels=['0-5%', '5-10%', '10-15%', '>15%'])

discount_analysis = df.groupby('discount_bucket').agg({
    'profit_margin_pct': 'mean',
    'revenue': 'sum',
    'profit': 'sum',
    'order_id': 'count'
}).round(2)

discount_analysis.columns = ['Avg_Profit_Margin_%', 'Total_Revenue', 'Total_Profit', 'Orders']
discount_analysis['Revenue_Per_Order'] = (discount_analysis['Total_Revenue'] / discount_analysis['Orders']).round(2)
discount_analysis = discount_analysis.reset_index()

discount_analysis.to_excel(writer, sheet_name='Discount Analysis', index=False)

# ============================================================================
# SHEET 6: LOSS-MAKING PRODUCTS
# ============================================================================
print("[6] Identifying Loss-Making Products...")

product_performance = df_detailed.groupby(['product_id', 'category', 'city']).agg({
    'selling_price': 'sum',
    'cost_price': 'sum',
    'profit': 'sum',
    'order_id': 'count'
}).reset_index()

product_performance.columns = ['Product_ID', 'Category', 'City', 'Revenue', 
                               'Cost', 'Profit', 'Orders']
product_performance['Profit_Margin_%'] = (product_performance['Profit'] / 
                                          product_performance['Revenue'] * 100).round(2)

loss_makers = product_performance[product_performance['Profit'] < 0].sort_values('Profit')
loss_makers.to_excel(writer, sheet_name='Loss-Making Products', index=False)

# ============================================================================
# SHEET 7: DELIVERY PERFORMANCE
# ============================================================================
print("[7] Creating Delivery Performance Report...")

delivery_perf = df.groupby(['city', 'delivery_sla_breach']).agg({
    'order_id': 'count',
    'delivery_time_minutes': 'mean',
    'repeat_customer_flag': 'mean'
}).round(2)

delivery_perf.columns = ['Orders', 'Avg_Delivery_Time', 'Repeat_Customer_Rate']
delivery_perf = delivery_perf.reset_index()
delivery_perf['SLA_Status'] = delivery_perf['delivery_sla_breach'].map({0: 'On-Time', 1: 'Delayed'})
delivery_perf = delivery_perf[['city', 'SLA_Status', 'Orders', 'Avg_Delivery_Time', 'Repeat_Customer_Rate']]

delivery_perf.to_excel(writer, sheet_name='Delivery Performance', index=False)

# ============================================================================
# SHEET 8: PEAK HOURS ANALYSIS
# ============================================================================
print("[8] Creating Peak Hours Analysis...")

df['hour'] = df['order_time'].str.split(':').str[0].astype(int)

hourly_analysis = df.groupby('hour').agg({
    'order_id': 'count',
    'delivery_time_minutes': 'mean',
    'delivery_sla_breach': 'mean',
    'revenue': 'sum'
}).round(2)

hourly_analysis.columns = ['Orders', 'Avg_Delivery_Time', 'SLA_Breach_Rate', 'Revenue']
hourly_analysis['Revenue_Per_Order'] = (hourly_analysis['Revenue'] / hourly_analysis['Orders']).round(2)
hourly_analysis = hourly_analysis.reset_index()

hourly_analysis.to_excel(writer, sheet_name='Peak Hours', index=False)

# Save Excel file
writer.close()

# ============================================================================
# APPLY FORMATTING (using openpyxl)
# ============================================================================
print("\n[9] Applying conditional formatting...")

from openpyxl import load_workbook

wb = load_workbook(excel_file)

# Format Executive Summary
ws = wb['Executive Summary']
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 20

# Header formatting
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Format City Performance with conditional formatting
ws = wb['City Performance']

# Set column widths
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws.column_dimensions[col].width = 15

# Apply headers
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Profit Margin heatmap (Column H)
ws.conditional_formatting.add(f'H2:H{len(city_perf)+1}',
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B'))

# Revenue bar chart (Column B)
ws.conditional_formatting.add(f'B2:B{len(city_perf)+1}',
    DataBarRule(start_type='min', start_value=0, end_type='max',
                color="5A8AC6", showValue=True))

# Format Category Analysis
ws = wb['Category Analysis']
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 20

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Profit margin color scale
ws.conditional_formatting.add(f'D2:D{len(category_perf)+1}',
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B'))

# Format Loss-Making Products - highlight negative profits
ws = wb['Loss-Making Products']
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws.column_dimensions[col].width = 15

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Highlight negative profits in red
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
red_font = Font(color="9C0006", bold=True)

for row in ws.iter_rows(min_row=2, max_row=len(loss_makers)+1, min_col=6, max_col=6):
    for cell in row:
        cell.fill = red_fill
        cell.font = red_font

wb.save(excel_file)

print(f"\n✓ Excel report saved: {excel_file}")
print("\nReport includes:")
print("  • Executive Summary (KPIs)")
print("  • City Performance (with heatmaps)")
print("  • Category Analysis")
print("  • Monthly Trends")
print("  • Discount Analysis")
print("  • Loss-Making Products")
print("  • Delivery Performance")
print("  • Peak Hours Analysis")

print("\n" + "="*70)
print("EXCEL REPORT GENERATION COMPLETE")
print("="*70)